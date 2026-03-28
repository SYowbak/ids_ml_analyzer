"""
IDS ML Analyzer - Professional Report Generator

Генерація професійних звітів у форматах CSV, Excel, PDF.
Підтримує українську мову та візуалізації.
"""

import pandas as pd
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    Image, PageBreak, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Професійний генератор звітів для IDS системи.
    Підтримує CSV, Excel (з графіками), PDF (з українською мовою).
    """
    
    SEVERITY_COLORS = {
        'critical': 'DC2626',
        'high': 'EF4444',
        'medium': 'F59E0B',
        'low': '10B981',
        'normal': '6366F1'
    }
    
    def __init__(self):
        self._setup_fonts()
    
    def _setup_fonts(self):
        """Налаштовує шрифти для PDF з підтримкою кирилиці."""
        self.font_name = 'Helvetica'  # Безпечний default
        
        # Список можливих шляхів до шрифтів
        font_paths = [
            Path(__file__).parent.parent.parent / 'assets' / 'fonts' / 'DejaVuSans.ttf',
            Path('C:/Windows/Fonts/arial.ttf'),
            Path('C:/Windows/Fonts/tahoma.ttf'),
            Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
            Path('/usr/share/fonts/TTF/DejaVuSans.ttf'),
        ]
        
        font_names = ['DejaVu', 'Arial', 'Tahoma', 'DejaVuSans']
        
        for font_path, font_name in zip(font_paths, font_names):
            try:
                if font_path.exists():
                    pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
                    self.font_name = font_name
                    logger.info(f"Шрифт {font_name} успішно зареєстровано: {font_path}")
                    return
            except Exception as e:
                logger.warning(f"Не вдалося зареєструвати шрифт {font_path}: {e}")
        
        logger.warning("Не знайдено жодного кириличного шрифту, використовуємо Helvetica (обмежена підтримка)")
    
    def export_csv(self, df: pd.DataFrame) -> bytes:
        """Експорт DataFrame у CSV з UTF-8 BOM для Excel."""
        output = io.BytesIO()
        output.write(b'\xef\xbb\xbf')
        output.write(df.to_csv(index=False).encode('utf-8'))
        return output.getvalue()

    def export_excel(self, df: pd.DataFrame, 
                     anomalies_df: Optional[pd.DataFrame] = None,
                     summary: Optional[dict] = None) -> bytes:
        """
        Експорт у професійний Excel з множинними аркушами.
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Всі записи', index=False)
            
            if anomalies_df is not None and not anomalies_df.empty:
                anomalies_df.to_excel(writer, sheet_name='Аномалії', index=False)
            
            if summary:
                self._create_summary_sheet(writer, summary)
            
            self._apply_excel_styling(writer)
        
        return output.getvalue()

    def _create_summary_sheet(self, writer, summary: dict):
        """Створює аркуш Summary з метриками."""
        summary_data = {
            'Метрика': [
                'Дата сканування',
                'Файл',
                'Модель',
                'Всього записів',
                'Виявлено аномалій',
                'Рівень ризику',
                'Безпечний трафік'
            ],
            'Значення': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                summary.get('filename', 'N/A'),
                summary.get('model_name', 'N/A'),
                summary.get('total', 0),
                summary.get('anomalies', 0),
                f"{summary.get('risk_score', 0)}%",
                f"{100 - summary.get('risk_score', 0)}%"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Підсумок', index=False)

    def _apply_excel_styling(self, writer):
        """Застосовує професійний стиль до Excel."""
        workbook = writer.book
        
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        
        danger_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

    def generate_pdf_report(self, summary: dict, 
                            details_df: Optional[pd.DataFrame] = None,
                            ai_analysis: Optional[str] = None,
                            executive_summary: Optional[str] = None) -> bytes:
        """
        Генерує професійний PDF звіт.
        """
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = self._create_pdf_styles()
        story = []
        
        story.extend(self._create_title_page(summary, styles))
        story.append(PageBreak())
        
        if executive_summary:
            story.extend(self._create_executive_section(executive_summary, styles))
            story.append(Spacer(1, 1*cm))
        
        story.extend(self._create_metrics_section(summary, styles))
        story.append(Spacer(1, 1*cm))
        
        if details_df is not None and not details_df.empty:
            story.extend(self._create_threats_table(details_df, styles))
            story.append(Spacer(1, 1*cm))
            
            # P3: Add Suspicious IP Report if IP context exists
            if 'src_ip' in details_df.columns or 'dst_ip' in details_df.columns:
                story.extend(self._create_suspicious_ips_section(details_df, styles))
                story.append(Spacer(1, 1*cm))
        
        if ai_analysis:
            story.extend(self._create_ai_section(ai_analysis, styles))
            story.append(Spacer(1, 1*cm))
        
        story.extend(self._create_recommendations_section(summary, styles))
        
        doc.build(story)
        return output.getvalue()

    def _create_pdf_styles(self) -> dict:
        """Створює стилі для PDF."""
        styles = getSampleStyleSheet()
        
        styles.add(ParagraphStyle(
            name='MainTitle',
            fontName=self.font_name,
            fontSize=24,
            alignment=TA_CENTER,
            spaceAfter=12,
            textColor=colors.HexColor('#1a1a2e')
        ))
        
        styles.add(ParagraphStyle(
            name='Subtitle',
            fontName=self.font_name,
            fontSize=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=30
        ))
        
        styles.add(ParagraphStyle(
            name='SectionHeader',
            fontName=self.font_name,
            fontSize=14,
            textColor=colors.HexColor('#6366f1'),
            spaceBefore=20,
            spaceAfter=10,
            borderPadding=5
        ))
        
        styles.add(ParagraphStyle(
            name='CustomBodyText',
            fontName=self.font_name,
            fontSize=10,
            alignment=TA_JUSTIFY,
            textColor=colors.HexColor('#334155'),
            spaceAfter=8,
            leading=14
        ))
        
        styles.add(ParagraphStyle(
            name='MetricValue',
            fontName=self.font_name,
            fontSize=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#6366f1')
        ))
        
        styles.add(ParagraphStyle(
            name='MetricLabel',
            fontName=self.font_name,
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b')
        ))
        
        return styles

    def _create_title_page(self, summary: dict, styles) -> list:
        """Створює титульну сторінку."""
        elements = []
        
        elements.append(Spacer(1, 4*cm))
        
        elements.append(Paragraph("IDS Security Report", styles['MainTitle']))
        elements.append(Paragraph(
            "Intrusion Detection System - Threat Analysis Report",
            styles['Subtitle']
        ))
        
        elements.append(Spacer(1, 2*cm))
        
        elements.append(HRFlowable(
            width="80%", thickness=1, 
            color=colors.HexColor('#6366f1'),
            spaceBefore=10, spaceAfter=10
        ))
        
        info_data = [
            ['Scan Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['File:', summary.get('filename', 'N/A')],
            ['Model:', summary.get('model_name', 'N/A')],
            ['Total Records:', f"{summary.get('total', 0):,}"],
            ['Threats Found:', f"{summary.get('anomalies', 0):,}"],
            ['Risk Score:', f"{summary.get('risk_score', 0)}%"]
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(info_table)
        
        elements.append(Spacer(1, 3*cm))
        
        risk = summary.get('risk_score', 0)
        if risk < 25:
            status_text = "LOW RISK"
            status_color = '#10b981'
        elif risk < 50:
            status_text = "MEDIUM RISK"
            status_color = '#f59e0b'
        elif risk < 75:
            status_text = "HIGH RISK"
            status_color = '#ef4444'
        else:
            status_text = "CRITICAL RISK"
            status_color = '#dc2626'
        
        status_style = ParagraphStyle(
            'StatusBadge',
            fontName=self.font_name,
            fontSize=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor(status_color)
        )
        elements.append(Paragraph(f"<b>{status_text}</b>", status_style))
        
        return elements

    def _create_executive_section(self, executive_summary: str, styles) -> list:
        """Створює секцію Executive Summary."""
        elements = []
        
        elements.append(Paragraph("Executive Summary", styles['SectionHeader']))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=10
        ))
        
        for line in executive_summary.split('\n'):
            if line.strip():
                if line.startswith('##'):
                    elements.append(Paragraph(
                        f"<b>{line.replace('#', '').strip()}</b>",
                        styles['CustomBodyText']
                    ))
                elif line.startswith('-') or line.startswith('*'):
                    elements.append(Paragraph(
                        f"• {line.lstrip('-* ').strip()}",
                        styles['CustomBodyText']
                    ))
                elif line[0].isdigit() and '.' in line[:3]:
                    elements.append(Paragraph(line.strip(), styles['CustomBodyText']))
                else:
                    elements.append(Paragraph(line.strip(), styles['CustomBodyText']))
        
        return elements

    def _create_metrics_section(self, summary: dict, styles) -> list:
        """Створює секцію з метриками."""
        elements = []
        
        elements.append(Paragraph("Scan Metrics", styles['SectionHeader']))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=15
        ))
        
        total = summary.get('total', 0)
        anomalies = summary.get('anomalies', 0)
        risk = summary.get('risk_score', 0)
        safe_pct = round((1 - anomalies/max(total, 1)) * 100, 1)
        
        metrics_data = [
            [f"{total:,}", f"{anomalies:,}", f"{risk}%", f"{safe_pct}%"],
            ['Total Records', 'Threats Found', 'Risk Score', 'Safe Traffic']
        ]
        
        metrics_table = Table(metrics_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        metrics_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.font_name),
            ('FONTNAME', (0, 1), (-1, 1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 18),
            ('FONTSIZE', (0, 1), (-1, 1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (2, 0), (2, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (3, 0), (3, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#64748b')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ]))
        
        elements.append(metrics_table)
        
        return elements

    def _create_threats_table(self, df: pd.DataFrame, styles) -> list:
        """Створює таблицю з топ загрозами."""
        elements = []
        
        elements.append(Paragraph("Top Detected Threats", styles['SectionHeader']))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=15
        ))
        
        if 'prediction' in df.columns:
            threat_counts = df['prediction'].value_counts().head(10)
            
            table_data = [['Threat Type', 'Count', 'Percentage']]
            total = len(df)
            
            for threat, count in threat_counts.items():
                pct = f"{(count/total)*100:.1f}%"
                table_data.append([str(threat), str(count), pct])
            
            threats_table = Table(table_data, colWidths=[8*cm, 3*cm, 3*cm])
            threats_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), self.font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(threats_table)
        
        return elements

    def _create_ai_section(self, ai_analysis: str, styles) -> list:
        """Створює секцію AI-аналізу."""
        elements = []
        
        elements.append(Paragraph("AI Threat Analysis", styles['SectionHeader']))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=15
        ))
        
        for line in ai_analysis.split('\n'):
            if line.strip():
                clean_line = line.replace('#', '').strip()
                if line.startswith('###') or line.startswith('##'):
                    elements.append(Paragraph(
                        f"<b>{clean_line}</b>",
                        styles['CustomBodyText']
                    ))
                elif line.startswith('-') or line.startswith('*'):
                    elements.append(Paragraph(
                        f"• {line.lstrip('-* ').strip()}",
                        styles['CustomBodyText']
                    ))
                else:
                    elements.append(Paragraph(clean_line, styles['CustomBodyText']))
        
        return elements

    def _create_recommendations_section(self, summary: dict, styles) -> list:
        """Створює секцію рекомендацій."""
        elements = []
        
        elements.append(Paragraph("Recommendations", styles['SectionHeader']))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=15
        ))
        
        risk = summary.get('risk_score', 0)
        
        if risk < 25:
            recs = [
                "Continue regular monitoring of network traffic",
                "Maintain current security policies",
                "Schedule periodic security reviews"
            ]
        elif risk < 50:
            recs = [
                "Review detected anomalies for false positives",
                "Update firewall rules based on findings",
                "Enable enhanced logging for suspicious sources",
                "Schedule security team review within 48 hours"
            ]
        elif risk < 75:
            recs = [
                "Immediately investigate high-severity threats",
                "Block suspicious IP addresses at firewall",
                "Check for lateral movement indicators",
                "Enable incident response protocols",
                "Notify security team and management"
            ]
        else:
            recs = [
                "CRITICAL: Initiate incident response immediately",
                "Isolate affected network segments",
                "Engage incident response team",
                "Preserve forensic evidence",
                "Prepare executive notification",
                "Consider external security support"
            ]
        
        for i, rec in enumerate(recs, 1):
            elements.append(Paragraph(f"{i}. {rec}", styles['CustomBodyText']))
        
        return elements


if __name__ == "__main__":
    print("ReportGenerator loaded successfully.")
